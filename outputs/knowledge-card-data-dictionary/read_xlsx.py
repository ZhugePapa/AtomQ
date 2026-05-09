#!/usr/bin/env python3
"""Read and print all content from an Excel xlsx file."""

import sys, os

def read_xlsx(filepath):
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
        import openpyxl

    wb = openpyxl.load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"  Sheet: {sheet_name}")
        print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
        print(f"{'='*80}\n")

        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
            rows.append([str(c) if c is not None else '' for c in row])

        if not rows:
            print("  (empty sheet)")
            continue

        col_widths = []
        for ci in range(len(rows[0])):
            mw = max(len(r[ci]) for r in rows if ci < len(r))
            col_widths.append(min(mw + 2, 50))

        total_w = sum(col_widths) + len(col_widths) * 3 + 1
        sep = '-' * min(total_w, 200)

        for i, row in enumerate(rows):
            cells = []
            for j, cell in enumerate(row):
                if j < len(col_widths):
                    d = cell[:47] + '...' if len(cell) > 50 else cell
                    cells.append(d.ljust(col_widths[j]))
            print(' | ' + ' | '.join(cells) + ' |')
            if i == 0:
                print(sep)

        print(sep)

if __name__ == '__main__':
    target = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        '软考刷题App_知识卡片数据字典.xlsx'
    )
    if not os.path.exists(target):
        print(f"ERROR: {target} not found")
        sys.exit(1)
    read_xlsx(target)
